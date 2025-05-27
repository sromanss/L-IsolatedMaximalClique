import time
import csv
import networkx as nx
import os
import random
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook
import algoritmi as alg

#**********************************************************************************************************
# * Funzione per generare un grafo secondo il modello G_{n,f,p} *
#**********************************************************************************************************
def generate_feature_graph(n, f, p):
    """
    Genera un grafo secondo il modello G_{n,f,p}
    n: numero di nodi
    f: numero di features
    p: probabilità che un nodo abbia una feature
    """
    G = nx.Graph()
    G.add_nodes_from(range(n))
    
    # Genera le feature per ogni nodo
    node_features = {i: set() for i in range(n)}
    for node in range(n):
        for feature in range(f):
            # genera un numero casuale tra 0 e 1; se è minore di p, aggiungi la feature al nodo
            if random.random() < p:
                node_features[node].add(feature)
    
    # Crea cliques tra nodi che condividono features
    for feature in range(f):
        # Trova tutti i nodi che hanno questa feature
        nodes_with_feature = [node for node, features in node_features.items() if feature in features]
        # Crea una clique tra questi nodi
        for i in range(len(nodes_with_feature)):
            for j in range(i+1, len(nodes_with_feature)):
                G.add_edge(nodes_with_feature[i], nodes_with_feature[j])
    
    return G
#**********************************************************************************************************
# * Funzione per analizzare il grafo *
#**********************************************************************************************************
def analyze_graph(G):
    """
    Calcola diverse metriche del grafo
    """
    metrics = {
        'nodes': G.number_of_nodes(),
        'edges': G.number_of_edges(),
        'density': nx.density(G)
        
    }
    return metrics
#**********************************************************************************************************
# * Funzione per eseguire i test su un grafo generato *
#**********************************************************************************************************

def esegui_test_generato(n, f, p, valori_L, euristiche, ripetizioni, cartella_output="risultati_generati"):
    """
    Esegue i test su un grafo generato secondo il modello G_{n,f,p}
    """
    # Crea la cartella di output se non esiste
    if not os.path.exists(cartella_output):
        os.makedirs(cartella_output)
        
    # Genera il grafo
    print(f"Generazione del grafo con parametri: n={n}, f={f}, p={p}")
    G = generate_feature_graph(n, f, p)
    metrics = analyze_graph(G)
    print(f"Grafo generato: nodi={metrics['nodes']}, archi={metrics['edges']}, densità={metrics['density']:.4f}")
    
    # Nome grafo per identificazione
    nome_grafo = f"G_{n}_{f}_{p}"
    
    # Prepara il file CSV per i risultati
    nome_file_csv = os.path.join(cartella_output, f"risultati_test_{nome_grafo}.csv")
    with open(nome_file_csv, 'w', newline='') as file_csv:
        writer = csv.writer(file_csv)
        writer.writerow(["L", "Euristica", "Tempo_Medio", "Chiamate_Ricorsive_Medie", "Clique_Trovate"])
        
        # Esegui i test per ogni combinazione di parametri
        for L in valori_L:
            for euristica in euristiche:
                print(f"Test con L={L}, euristica={euristica}")
                
                # Liste per memorizzare i risultati
                tempi = []
                chiamate_ricorsive = []
                conteggio_clique = 0
                
                # Esegui più volte per calcolare la media
                for i in range(ripetizioni):
                    
                    # Misura il tempo di esecuzione
                    inizio = time.time()
                    
                    if euristica == -1:
                        # Numero totale di clique massimali (Bron-Kerbosch)
                        cliques, num_chiamate = alg.trova_clique_massimali3(G)    
                    elif euristica == 0:
                        # Approccio trova e filtra
                        cliques, num_chiamate = alg.trova_clique_massimali3(G)
                        clique_isolate = alg.filtra_clique_isolate(G, cliques, L)  
                    else:
                        # Approccio con euristiche
                        clique_isolate, num_chiamate = alg.trova_clique_massimali_L_isolated2(G, L, euristica)
                    
                    # Registra il tempo
                    fine = time.time()
                    tempo_esecuzione = fine - inizio
                    
                    # Memorizza i risultati
                    tempi.append(tempo_esecuzione)
                    chiamate_ricorsive.append(num_chiamate)
                    
                    # Usa l'ultimo valore per il conteggio delle clique
                    if i == ripetizioni - 1:
                        if euristica == -1:
                            conteggio_clique = len(cliques)
                        else:
                            conteggio_clique = len(clique_isolate)
                
                # Calcola le medie
                tempo_medio = sum(tempi) / len(tempi)
                chiamate_medie = sum(chiamate_ricorsive) / len(chiamate_ricorsive)
                
                # Salva i risultati nel CSV
                writer.writerow([L, euristica, tempo_medio, chiamate_medie, conteggio_clique])
                
                print(f"  Risultati: Tempo medio: {tempo_medio:.4f}s, Chiamate medie: {chiamate_medie:.2f}, Clique trovate: {conteggio_clique}")
    
    print(f"Risultati salvati in {nome_file_csv}")
    
    # Genera i grafici
    crea_grafici(nome_file_csv, cartella_output, nome_grafo, G, metrics)

# **********************************************************************************************************
# * Funzione per la creazione dei grafici e salvataggio in Excel *
# **********************************************************************************************************

def crea_grafici(file_csv, cartella_output, nome_grafo, G, metrics):
    """
    Crea il file Excel con i risultati del test
    """
    excel_path = os.path.join(cartella_output, "TempiEsecuzione_Generati.xlsx")
    
    # Verifica se il file Excel esiste già
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws_data = wb.active
        # Aggiungi sempre in fondo al file
        row_start = ws_data.max_row + 2
    else:
        # Crea nuovo file Excel
        wb = Workbook()
        ws_data = wb.active
        ws_data.title = "Dati"
        row_start = 1
    
    # Leggi i dati dal CSV
    L_values = []
    euristiche = []
    tempi = []
    chiamate = []
    clique = []

    with open(file_csv, 'r') as f:
        reader = csv.reader(f)
        next(reader)  # Salta l'intestazione
        for row_data in reader:
            L_values.append(float(row_data[0]))
            euristiche.append(int(row_data[1]))
            tempi.append(float(row_data[2]))
            chiamate.append(float(row_data[3]))
            clique.append(int(row_data[4]))

    # Organizza i dati per L e euristica
    L_unici = sorted(list(set(L_values)))
    euristiche_uniche = sorted(list(set(euristiche)))
    dati_organizzati = {L: {e: {'tempo': 0, 'chiamate': 0, 'clique': 0} 
                           for e in euristiche_uniche} for L in L_unici}
    
    for i in range(len(L_values)):
        L = L_values[i]
        e = euristiche[i]
        dati_organizzati[L][e]['tempo'] = tempi[i]
        dati_organizzati[L][e]['chiamate'] = chiamate[i]
        dati_organizzati[L][e]['clique'] = clique[i]

    # Scrivi intestazione con info grafo
    ws_data[f'A{row_start}'] = f"GRAFO GENERATO {nome_grafo} CON N={metrics['nodes']}, M={metrics['edges']}, DENSITÀ={metrics['density']:.4f}"
    ws_data[f'A{row_start}'].font = Font(bold=True, color="008000")  # Verde (codice RGB)
    ws_data.merge_cells(f'A{row_start}:H{row_start}')

    # Intestazioni prima tabella (Tempi)
    ws_data[f'A{row_start+1}'] = 'L'
    for i, e in enumerate(euristiche_uniche):
        ws_data[f'{chr(66+i)}{row_start+1}'] = f'Euristica {e}'
    
    # Applica il grassetto alle intestazioni
    for col in range(1, 2+len(euristiche_uniche)):
        ws_data[f'{chr(64+col)}{row_start+1}'].font = Font(bold=True)

    # Riempi la prima tabella (Tempi)
    row = row_start + 2
    for L in L_unici:
        ws_data[f'A{row}'] = L
        for i, e in enumerate(euristiche_uniche):
            ws_data[f'{chr(66+i)}{row}'] = dati_organizzati[L][e]['tempo']
        row += 1

    # Intestazioni seconda tabella (Chiamate ricorsive)
    col_offset = 8
    ws_data[f'{chr(65+col_offset)}{row_start+1}'] = 'L'
    for i, e in enumerate(euristiche_uniche):
        ws_data[f'{chr(65+col_offset+1+i)}{row_start+1}'] = f'Euristica {e}'

    # Applica il grassetto alle intestazioni della seconda tabella
    for col in range(col_offset, col_offset+len(euristiche_uniche)+1):
        ws_data[f'{chr(65+col)}{row_start+1}'].font = Font(bold=True)

    # Riempi la seconda tabella (Chiamate ricorsive)
    row = row_start + 2
    for L in L_unici:
        ws_data[f'{chr(65+col_offset)}{row}'] = L
        for i, e in enumerate(euristiche_uniche):
            ws_data[f'{chr(65+col_offset+1+i)}{row}'] = dati_organizzati[L][e]['chiamate']
        row += 1
    
    # Intestazioni terza tabella (Clique trovate)
    col_offset = 16  # Aumentato offset per la nuova tabella
    ws_data[f'{chr(65+col_offset)}{row_start+1}'] = 'L'
    for i, e in enumerate(euristiche_uniche):
        ws_data[f'{chr(65+col_offset+1+i)}{row_start+1}'] = f'Euristica {e}'

    # Applica il grassetto alle intestazioni della terza tabella
    for col in range(col_offset, col_offset+len(euristiche_uniche)+1):
        ws_data[f'{chr(65+col)}{row_start+1}'].font = Font(bold=True)

    # Riempi la terza tabella (Clique trovate)
    row = row_start + 2
    for L in L_unici:
        ws_data[f'{chr(65+col_offset)}{row}'] = L
        for i, e in enumerate(euristiche_uniche):
            ws_data[f'{chr(65+col_offset+1+i)}{row}'] = dati_organizzati[L][e]['clique']
        row += 1
    
    # Salva il file Excel
    try:
        wb.save(excel_path)
        print(f"\nRisultati aggiunti in Excel: {excel_path}")
    except PermissionError:
        print(f"\nERRORE: Impossibile salvare il file {excel_path}")
        print("Chiudi il file Excel e riprova.")

    # Stampa il numero di clique trovate
    print("\nNumero di clique L-isolated trovate:")
    print("-" * 50)
    print("L\tEuristica\tClique trovate")
    print("-" * 50)
    for i in range(len(L_values)):
        print(f"{L_values[i]}\t{euristiche[i]}\t\t{clique[i]}")
    print("-" * 50)
    
    print(f"\nRisultati salvati in Excel: {excel_path}")

# **********************************************************************************************************
# * Funzione principale per l'analisi sperimentale *
# ********************************************************************************************************
def run_experimental_analysis(cartella_output="risultati_sperimentali"):
    """
    Esegue l'analisi sperimentale completa
    """
    if not os.path.exists(cartella_output):
        os.makedirs(cartella_output)
    
    # Definizione dei valori per ogni parametro
    n_values = [10, 60, 120, 240, 440]              # Numero di nodi
    f_values = [5, 10, 20, 40, 80]                      # Numero di features
    p_values = [0.025, 0.050, 0.075, 0.100, 0.125, 0.150, 0.175, 0.225]      # Probabilità che un nodo n abbia la feature f
    l_values = [20, 40, 60, 80, 100]                         # Valori di L da testare
    euristiche = [-1, 0, 1, 2, 3, 4]                        # Euristiche da testare
    ripetizioni = 5                                    # Numero di ripetizioni per calcolare la media
    
    # PRIMA COPPIA: -numero di cliques al variare di L (n fisso, f fisso, p fisso)
    #               -running time al variare di L (n fisso, f fisso, p fisso)

    # SECONDA COPPIA: -numero di cliques al variare di f (n fisso, p fisso, l fisso)
    #                 -running time al variare di f (n fisso, p fisso, l fisso)

    # TERZA COPPIA: -numero di cliques al variare di n (f fisso, p fisso, L fisso)
    #               -running time al variare di n (f fisso, p fisso, L fisso)

    # QUARTA COPPIA: -numero di cliques al variare di p (n fisso, f fisso, L fisso)
    #                -running time al variare di p (n fisso, f fisso, L fisso)
    
    # Test 1: Variazione di L (n, f, p fissi)
    print("\nTEST 1: Variazione di l (n=200, f=45, p=0.1)")
    try:
        esegui_test_generato(
            n=200,          # fisso
            f=45,           # fisso
            p=0.1,        # fisso
            valori_L=l_values,  # variabile
            euristiche=euristiche,
            ripetizioni=ripetizioni,
            cartella_output=cartella_output
        )
    except Exception as e:
        print(f"Errore durante test 1: {str(e)}")

    # Test 2: Variazione di f (n, p, L fissi)
    print("\nTEST 2: Variazione di f (n=200, p=0.1, L=40)")
    for f in f_values:
        try:
            esegui_test_generato(
                n=200,          # fisso
                f=f,            # variabile
                p=0.1,        # fisso
                valori_L=[40],  # fisso
                euristiche=euristiche,
                ripetizioni=ripetizioni,
                cartella_output=cartella_output
            )
        except Exception as e:
            print(f"Errore durante test 2 con f={f}: {str(e)}")

    # Test 3: Variazione di n (f, p, L fissi)
    print("\nTEST 3: Variazione di n (f=45, p=0.1, L=40)")
    for n in n_values:
        try:
            esegui_test_generato(
                n=n,            # variabile
                f=45,           # fisso
                p=0.1,        # fisso
                valori_L=[40],  # fisso
                euristiche=euristiche,
                ripetizioni=ripetizioni,
                cartella_output=cartella_output
            )
        except Exception as e:
            print(f"Errore durante test 3 con n={n}: {str(e)}")

    
    # Test 4: Variazione di p (n, f, L fissi)
    print("\nTEST 4: Variazione di p (n=200, f=45, L=40)")
    for p in p_values:
        try:
            esegui_test_generato(
                n=200,          # fisso
                f=45,           # fisso
                p=p,            # variabile
                valori_L=[40],  # fisso
                euristiche=euristiche,
                ripetizioni=ripetizioni,
                cartella_output=cartella_output
            )
        except Exception as e:
            print(f"Errore durante test 4 con p={p}: {str(e)}")
        
if __name__ == "__main__":
    run_experimental_analysis()