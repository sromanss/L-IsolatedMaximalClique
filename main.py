import time
import csv
import networkx as nx
import os
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook
import algoritmi as alg

def esegui_test(percorso_grafo, valori_L, euristiche, ripetizioni, cartella_output="risultati"):
    
    # Crea la cartella di output se non esiste
    if not os.path.exists(cartella_output):
        os.makedirs(cartella_output)
    
    # Carica il grafo
    print(f"Caricamento del grafo: {percorso_grafo}")
    G = nx.read_edgelist(percorso_grafo, comments="%")
    print(f"Grafo caricato: nodi={len(G)}, archi={G.number_of_edges()}")
    
    # Prepara il file CSV per i risultati
    nome_file_csv = os.path.join(cartella_output, "risultati_test.csv")
    with open(nome_file_csv, 'w', newline='') as file_csv:
        writer = csv.writer(file_csv)
        writer.writerow(["L", "Euristica", "Tempo_Medio", "Chiamate_Ricorsive_Medie", "Clique_Trovate"])
        
        # Esegui i test per ogni combinazione di parametri
        for L in valori_L:
            for euristica in euristiche:
                print(f"Test con L={L}, euristica={euristica}")
                
                # Liste per memorizzare i risultati
                tempi = []
                chiamate_ricorsive = []
                conteggio_clique = 0
                
                # Esegui più volte per calcolare la media
                for i in range(ripetizioni):
                    
                    # Misura il tempo di esecuzione
                    inizio = time.time()
                    
                    if euristica == 0:
                        # Approccio trova e filtra
                        cliques, num_chiamate = alg.trova_clique_massimali2(G)
                        clique_isolate = alg.filtra_clique_isolate(G, cliques, L)
                    else:
                        # Approccio con euristiche
                        clique_isolate, num_chiamate = alg.trova_clique_massimali_L_isolated2(G, L, euristica)
                    
                    
                    # Registra il tempo
                    fine = time.time()
                    tempo_esecuzione = fine - inizio
                    
                    # Memorizza i risultati
                    tempi.append(tempo_esecuzione)
                    chiamate_ricorsive.append(num_chiamate)
                    
                    # Usa l'ultimo valore per il conteggio delle clique
                    if i == ripetizioni - 1:
                        conteggio_clique = len(clique_isolate)
                
                # Calcola le medie
                tempo_medio = sum(tempi) / len(tempi)
                chiamate_medie = sum(chiamate_ricorsive) / len(chiamate_ricorsive)
                
                # Salva i risultati nel CSV
                writer.writerow([L, euristica, tempo_medio, chiamate_medie, conteggio_clique])
                
                print(f"  Risultati: Tempo medio: {tempo_medio:.4f}s, Chiamate medie: {chiamate_medie:.2f}, Clique trovate: {conteggio_clique}")
    
    print(f"Risultati salvati in {nome_file_csv}")
    
    # Genera i grafici
    crea_grafici(nome_file_csv, cartella_output, percorso_grafo)

def get_graph_name(path):
    """Estrae il nome del grafo dal percorso"""
    filename = os.path.basename(path)
    if filename.startswith('out.'):
        return filename[4:]  # Rimuove 'out.' dal nome
    return filename

def crea_grafici(file_csv, cartella_output, percorso_grafo):
    """Crea il file Excel con i risultati del test"""
    excel_path = os.path.join(cartella_output, "TempiEsecuzione.xlsx")
    nome_grafo = get_graph_name(percorso_grafo)
    
    # Verifica se il file Excel esiste già
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws_data = wb.active
        # Aggiungi sempre in fondo al file
        row_start = ws_data.max_row + 2
    else:
        # Crea nuovo file Excel
        wb = Workbook()
        ws_data = wb.active
        ws_data.title = "Dati"
        row_start = 1
    
    # Leggi i dati dal CSV
    L_values = []
    euristiche = []
    tempi = []
    chiamate = []
    clique = []

    with open(file_csv, 'r') as f:
        reader = csv.reader(f)
        next(reader)  # Salta l'intestazione
        for row_data in reader:
            L_values.append(float(row_data[0]))
            euristiche.append(int(row_data[1]))
            tempi.append(float(row_data[2]))
            chiamate.append(float(row_data[3]))
            clique.append(int(row_data[4]))

    # Organizza i dati per L e euristica
    L_unici = sorted(list(set(L_values)))
    euristiche_uniche = sorted(list(set(euristiche)))
    dati_organizzati = {L: {e: {'tempo': 0, 'chiamate': 0, 'clique': 0} 
                           for e in euristiche_uniche} for L in L_unici}
    
    for i in range(len(L_values)):
        L = L_values[i]
        e = euristiche[i]
        dati_organizzati[L][e]['tempo'] = tempi[i]
        dati_organizzati[L][e]['chiamate'] = chiamate[i]
        dati_organizzati[L][e]['clique'] = clique[i]

    # Scrivi intestazione con info grafo
    G = nx.read_edgelist(percorso_grafo, comments="%")
    ws_data[f'A{row_start}'] = f"GRAFO CON N={len(G)} E M={G.number_of_edges()}; Nome networkx: {nome_grafo}"
    ws_data[f'A{row_start}'].font = Font(bold=True, color="008000")  # Verde (codice RGB), in modo da riconoscere i nuovi dati da quelli vecchi
    ws_data.merge_cells(f'A{row_start}:H{row_start}')

    # Intestazioni prima tabella (Tempi)
    ws_data[f'A{row_start+1}'] = 'L'
    ws_data[f'B{row_start+1}'] = 'Euristica 0'
    ws_data[f'C{row_start+1}'] = 'Euristica 1'
    ws_data[f'D{row_start+1}'] = 'Euristica 2'
    ws_data[f'E{row_start+1}'] = 'Euristica 3'
    ws_data[f'F{row_start+1}'] = 'Euristica 4'
    
    # Applica il grassetto alle intestazioni
    for col in range(1, 7):
        ws_data[f'{chr(64+col)}{row_start+1}'].font = Font(bold=True)

    # Riempi la prima tabella (Tempi)
    row = row_start + 2
    for L in L_unici:
        ws_data[f'A{row}'] = L
        for i, e in enumerate(euristiche_uniche):
            ws_data[f'{chr(65+i)}{row}'] = dati_organizzati[L][e]['tempo']
        row += 1

    # Intestazioni seconda tabella (Chiamate ricorsive)
    col_offset = 8
    ws_data[f'{chr(65+col_offset)}{row_start+1}'] = 'L'
    for i, e in enumerate(euristiche_uniche):
        ws_data[f'{chr(65+col_offset+1+i)}{row_start+1}'] = f'Euristica {e}'

    # Applica il grassetto alle intestazioni della seconda tabella
    for col in range(col_offset, col_offset+len(euristiche_uniche)+1):
        ws_data[f'{chr(65+col)}{row_start+1}'].font = Font(bold=True)

    # Riempi la seconda tabella (Chiamate ricorsive)
    row = row_start + 2
    for L in L_unici:
        ws_data[f'{chr(65+col_offset)}{row}'] = L
        for i, e in enumerate(euristiche_uniche):
            ws_data[f'{chr(65+col_offset+i)}{row}'] = dati_organizzati[L][e]['chiamate']
        row += 1
    
     # Intestazioni terza tabella (Clique trovate)
    col_offset = 16  # Aumentato offset per la nuova tabella
    ws_data[f'{chr(65+col_offset)}{row_start+1}'] = 'L'
    for i, e in enumerate(euristiche_uniche):
        ws_data[f'{chr(65+col_offset+1+i)}{row_start+1}'] = f'Euristica {e}'

    # Applica il grassetto alle intestazioni della terza tabella
    for col in range(col_offset, col_offset+len(euristiche_uniche)+1):
        ws_data[f'{chr(65+col)}{row_start+1}'].font = Font(bold=True)

    # Riempi la terza tabella (Clique trovate)
    row = row_start + 2
    for L in L_unici:
        ws_data[f'{chr(65+col_offset)}{row}'] = L
        for i, e in enumerate(euristiche_uniche):
            ws_data[f'{chr(65+col_offset+1+i)}{row}'] = dati_organizzati[L][e]['clique']
        row += 1
    
    # Salva il file Excel
    try:
        wb.save(excel_path)
        print(f"\nRisultati aggiunti in Excel: {excel_path}")
    except PermissionError:
        print(f"\nERRORE: Impossibile salvare il file {excel_path}")
        print("Chiudi il file Excel e riprova.")

    # Stampa il numero di clique trovate
    print("\nNumero di clique L-isolated trovate:")
    print("-" * 50)
    print("L\tEuristica\tClique trovate")
    print("-" * 50)
    for i in range(len(L_values)):
        print(f"{L_values[i]}\t{euristiche[i]}\t\t{clique[i]}")
    print("-" * 50)
    
    print(f"\nRisultati salvati in Excel: {excel_path}")

if __name__ == "__main__":
    
    # Parametri da testare
    grafi = [
        #"C:/Users/simon/Downloads/out.ucidata-zachary",                         #n=34, m=78
        #"C:/Users/simon/Downloads/out.subelj_euroroad_euroroad",                #n=1174, m=1417
        #"C:/Users/simon/Downloads/out.petster-hamster-household",               #n=921, m=4032
        #"C:/Users/simon/Downloads/out.opsahl-powergrid",                        #n=4941, m=6594
        #"C:/Users/simon/Downloads/out.loc-brightkite_edges",                    #n=58228, m=214078
        #"C:/Users/simon/Downloads/out.petster-cat-household",                   #n=105138, m=494858
        #"C:/Users/simon/Downloads/out.loc-gowalla_edges",                       #196591, m=950327
        #"C:/Users/simon/Downloads/out.livemocha",                               #n=104103, m=2193083
        "C:/Users/simon/Downloads/out.petster-dog-household",                   #n=260390, m=2148179
        "C:/Users/simon/Downloads/out.roadNet-CA"                              #n=1965206, m=2766366
    ]
    valori_L = list(range(1, 11)) + [20, 30]   # Valori di L da testare
    euristiche = [0, 1, 2, 3, 4]
    
    # Ciclo su tutti i grafi
    for grafo in grafi:
        print("\n" + "="*80)
        print(f"Iniziando test su grafo: {get_graph_name(grafo)}")
        print("="*80 + "\n")
        
        try:
            # Esegui i test
            esegui_test(
                percorso_grafo=grafo,
                valori_L=valori_L,
                euristiche=euristiche,
                ripetizioni=10,
                cartella_output="risultati_test"
            )
        except Exception as e:
            print(f"Errore durante l'elaborazione del grafo {get_graph_name(grafo)}:")
            print(f"  {str(e)}")
            print("Passando al grafo successivo...\n")
            continue
            
        print(f"\nCompletato test su grafo: {get_graph_name(grafo)}\n")